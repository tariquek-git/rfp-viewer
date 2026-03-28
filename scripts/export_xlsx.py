#!/usr/bin/env python3
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = "public/rfp_data.json"
OUTPUT = "BSB_RFP_QA_Report.xlsx"

with open(DATA_FILE) as f:
    root = json.load(f)

questions = root['questions']
wb = Workbook()

FONT = Font(name='Arial', size=10)
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

CONF_MAP = {'GREEN': GREEN_FILL, 'YELLOW': YELLOW_FILL, 'RED': RED_FILL}

COLS = ['ref','category','number','topic','requirement','a_oob','b_config','c_custom','d_dnm',
        'compliant','confidence','bullet','paragraph','pricing','capability','availability',
        'strategic','notes','rationale','committee_review','committee_risk','committee_score']

HEADERS = ['Ref','Category','#','Topic','Requirement','A: OOB','B: Config','C: Custom','D: DNM',
           'Compliant','Confidence','Bullet Response','Paragraph Response','Pricing','Capability',
           'Availability','Strategic','Notes','Rationale','Committee Review','Committee Risk','Committee Score']

def style_header(ws, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"
    ws.freeze_panes = 'A2'

def write_questions(ws, qs):
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(HEADERS))

    for r, q in enumerate(qs, 2):
        for c, key in enumerate(COLS, 1):
            val = q.get(key, '')
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if key == 'confidence' and val in CONF_MAP:
                cell.fill = CONF_MAP[val]

    widths = {'ref':12,'category':20,'number':5,'topic':25,'requirement':40,'a_oob':6,'b_config':6,
              'c_custom':6,'d_dnm':6,'compliant':9,'confidence':11,'bullet':50,'paragraph':50,
              'pricing':15,'capability':12,'availability':14,'strategic':40,'notes':30,
              'rationale':35,'committee_review':40,'committee_risk':15,'committee_score':8}
    for c, key in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(key, 15)

# Sheet 1: All Questions
ws1 = wb.active
ws1.title = "All Questions"
write_questions(ws1, questions)

# Sheet 2: Summary by Category
ws2 = wb.create_sheet("Summary by Category")
sum_headers = ['Category','Total','GREEN','YELLOW','RED','% Compliant']
for i, h in enumerate(sum_headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, len(sum_headers))

cats = {}
for q in questions:
    cat = q.get('category', 'Unknown')
    if cat not in cats:
        cats[cat] = {'total': 0, 'GREEN': 0, 'YELLOW': 0, 'RED': 0, 'compliant': 0}
    cats[cat]['total'] += 1
    conf = q.get('confidence', '')
    if conf in cats[cat]:
        cats[cat][conf] += 1
    if q.get('compliant') == 'Y':
        cats[cat]['compliant'] += 1

for r, (cat, s) in enumerate(cats.items(), 2):
    ws2.cell(row=r, column=1, value=cat).font = FONT
    ws2.cell(row=r, column=2, value=s['total']).font = FONT
    g = ws2.cell(row=r, column=3, value=s['GREEN']); g.font = FONT; g.fill = GREEN_FILL
    y = ws2.cell(row=r, column=4, value=s['YELLOW']); y.font = FONT; y.fill = YELLOW_FILL
    rd = ws2.cell(row=r, column=5, value=s['RED']); rd.font = FONT; rd.fill = RED_FILL
    pct = s['compliant'] / s['total'] if s['total'] else 0
    p = ws2.cell(row=r, column=6, value=pct); p.font = FONT; p.number_format = '0.0%'
    for c in range(1, 7):
        ws2.cell(row=r, column=c).border = THIN_BORDER

for w, col in [(25,1),(8,2),(8,3),(8,4),(8,5),(12,6)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# Sheet 3: RED Questions
ws3 = wb.create_sheet("RED Questions")
reds = [q for q in questions if q.get('confidence') == 'RED']
write_questions(ws3, reds)

# Sheet 4: QA Changes Log
ws4 = wb.create_sheet("QA Changes Log")
log_headers = ['Question Ref', 'Field', 'Change Description', 'Type']
for i, h in enumerate(log_headers, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, len(log_headers))

changes = [
    ("Partner Relationships 1", "notes", "Renamed Aite-Novarica → Datos", "Name Fix"),
    ("Partner Relationships 7", "bullet", "Renamed Aite-Novarica → Datos", "Name Fix"),
    ("Partner Relationships 7", "paragraph", "Renamed Aite-Novarica → Datos", "Name Fix"),
    ("Partner Relationships 13", "bullet", "Softened deconversion claim: 'in production' → 'developed with'", "Hallucination Fix"),
    ("Partner Relationships 13", "paragraph", "Softened deconversion claim: 'in production' → 'developed with'", "Hallucination Fix"),
    ("Partner Relationships 14", "strategic", "Softened Mastercard EXCLUSIVE → strategic", "Hallucination Fix"),
    ("Technology 24", "bullet", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix"),
    ("Technology 24", "paragraph", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix"),
    ("Technology 24", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix"),
    ("Technology 24", "paragraph", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix"),
    ("Activation and Fulfillment 15", "rationale", "Removed self-referential audit notes", "Cleanup"),
    ("Compliance & Reporting 3", "strategic", "Renamed Aite-Novarica → Datos", "Name Fix"),
    ("Processing 19", "bullet", "Fixed Verafin: 'runs in production' → 'planned for integration'", "Hallucination Fix"),
    ("Processing 19", "paragraph", "Fixed Verafin: 'runs in production' → 'planned for integration'", "Hallucination Fix"),
    ("Processing 22", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix"),
    ("Processing 26", "rationale", "Removed self-referential audit notes", "Cleanup"),
    ("Processing 32", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix"),
    ("Processing 42", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix"),
    ("Accounting & Finance 3", "strategic", "Renamed Aite-Novarica → Datos", "Name Fix"),
    ("Accounting & Finance 4", "strategic", "Renamed Aite-Novarica → Datos", "Name Fix"),
    ("Accounting & Finance 15", "availability", "Updated past date Est. Q1 2026 → Est. Q2 2026", "Date Fix"),
    ("Product Operations 21", "paragraph", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix"),
    ("Product Operations 33", "paragraph", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix"),
    ("BULK (68 questions)", "bullet/paragraph", "Added terminal punctuation to 68 responses", "Formatting"),
]

# Previous cleanup passes
prev_changes = [
    ("All (20 questions)", "multiple", "Removed all Coast Capital Savings references (fabricated client)", "Hallucination Fix"),
    ("All", "multiple", "Removed all Continental Currency references (fabricated client)", "Hallucination Fix"),
    ("Activation and Fulfillment 15", "paragraph", "Removed fabricated 73% first-transaction metric", "Hallucination Fix"),
    ("Processing 26", "paragraph", "Removed fabricated 99.997% settlement accuracy metric", "Hallucination Fix"),
    ("Multiple", "multiple", "Replaced East Coast references with Eastern time zone", "Accuracy Fix"),
    ("Partner Relationships 8", "paragraph", "Fixed Wolfsberg FCCQ 45/45 → completed self-assessment", "Hallucination Fix"),
]

all_changes = prev_changes + changes
for r, (ref, field, desc, typ) in enumerate(all_changes, 2):
    ws4.cell(row=r, column=1, value=ref).font = FONT
    ws4.cell(row=r, column=2, value=field).font = FONT
    ws4.cell(row=r, column=3, value=desc).font = FONT
    c4 = ws4.cell(row=r, column=4, value=typ); c4.font = FONT
    if typ == 'Hallucination Fix':
        c4.fill = RED_FILL
    elif typ == 'Consistency Fix':
        c4.fill = YELLOW_FILL
    elif typ == 'Name Fix':
        c4.fill = PatternFill('solid', fgColor='BDD7EE')
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = THIN_BORDER
        ws4.cell(row=r, column=c).alignment = WRAP

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 55
ws4.column_dimensions['D'].width = 18

wb.save(OUTPUT)
print(f"Saved {OUTPUT}")
print(f"  Sheet 1: All Questions ({len(questions)} rows)")
print(f"  Sheet 2: Summary by Category ({len(cats)} categories)")
print(f"  Sheet 3: RED Questions ({len(reds)} rows)")
print(f"  Sheet 4: QA Changes Log ({len(all_changes)} entries)")
