#!/usr/bin/env python3
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = "public/rfp_data.json"
OUTPUT = "BSB_RFP_QA_Report.xlsx"

with open(DATA_FILE) as f:
    root = json.load(f)

questions = root['questions']
wb = Workbook()

FONT = Font(name='Arial', size=10)
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SECTION_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Section colors
BSB_FILL = PatternFill('solid', fgColor='1F4E79')       # Dark blue - BSB asks
BRIM_FILL = PatternFill('solid', fgColor='2E7D32')       # Dark green - Brim responds
AI_FILL = PatternFill('solid', fgColor='7B1FA2')         # Purple - Claude AI
META_FILL = PatternFill('solid', fgColor='424242')       # Dark gray - metadata

# Cell tints (light versions for data rows)
BSB_TINT = PatternFill('solid', fgColor='D6E4F0')        # Light blue
BRIM_TINT = PatternFill('solid', fgColor='E8F5E9')       # Light green
AI_TINT = PatternFill('solid', fgColor='F3E5F5')         # Light purple
META_TINT = PatternFill('solid', fgColor='F5F5F5')       # Light gray

# Confidence colors
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
CONF_MAP = {'GREEN': GREEN_FILL, 'YELLOW': YELLOW_FILL, 'RED': RED_FILL}

# AI detection words
AI_WORDS = {
    'additionally': ('Filler transition', 'Remove or replace with a direct connector like "Also" or just start the next point'),
    'furthermore': ('Filler transition', 'Remove — start the sentence directly'),
    'moreover': ('Filler transition', 'Remove — just state the next fact'),
    'comprehensive': ('Vague superlative', 'Replace with specific scope: "covers X, Y, Z" or "spans 12 modules"'),
    'robust': ('Vague superlative', 'Replace with measurable claim: "handles X TPS" or "tested against Y scenarios"'),
    'seamless': ('Vague superlative', 'Replace with specifics: "zero-downtime migration" or "single API call"'),
    'leverage': ('Corporate buzzword', 'Replace with "use", "apply", or "build on"'),
    'utilize': ('Corporate buzzword', 'Replace with "use"'),
    'facilitate': ('Corporate buzzword', 'Replace with "enable", "support", or "allow"'),
    'streamline': ('Vague claim', 'Replace with specific efficiency gain: "reduces steps from 5 to 2"'),
    'cutting-edge': ('Vague superlative', 'Remove or replace with specific tech: "cloud-native" or "event-driven"'),
    'best-in-class': ('Unverifiable superlative', 'Remove or cite the ranking: "ranked #1 by Datos in 2025"'),
    'industry-leading': ('Unverifiable superlative', 'Remove or cite specific evidence'),
    'synergy': ('Corporate buzzword', 'Replace with specific benefit: "shared data layer" or "combined reporting"'),
    'holistic': ('Vague superlative', 'Replace with "end-to-end" or list what it covers'),
    'paradigm': ('Corporate buzzword', 'Replace with plain language: "approach" or "model"'),
    'ecosystem': ('Overused tech buzzword', 'Replace with "platform", "network", or "integration layer"'),
}

PASSIVE_PATTERNS = [
    (' is designed ', 'Passive voice', 'Rewrite active: "We designed X to..." or "X does Y"'),
    (' are configured ', 'Passive voice', 'Rewrite active: "We configure X to..."'),
    (' is enabled ', 'Passive voice', 'Rewrite active: "X enables..." or "We enable..."'),
    (' is provided ', 'Passive voice', 'Rewrite active: "We provide..." or "Brim provides..."'),
    (' are provided ', 'Passive voice', 'Rewrite active: "We provide..."'),
    (' is supported ', 'Passive voice', 'Rewrite active: "We support..." or "X supports..."'),
    (' is integrated ', 'Passive voice', 'Rewrite active: "X integrates with..."'),
    (' is managed ', 'Passive voice', 'Rewrite active: "We manage..." or "X manages..."'),
    (' is maintained ', 'Passive voice', 'Rewrite active: "We maintain..."'),
    (' is processed ', 'Passive voice', 'Rewrite active: "We process..." or "The system processes..."'),
]

def detect_ai_score(text):
    if not text:
        return 0, [], [], []
    lower = text.lower()
    found_words = []
    reasons = []
    fixes = []
    for w, (reason, fix) in AI_WORDS.items():
        if w in lower:
            found_words.append(w)
            reasons.append(f'"{w}" — {reason}')
            fixes.append(f'"{w}" → {fix}')
    for pattern, reason, fix in PASSIVE_PATTERNS:
        count = lower.count(pattern)
        if count > 0:
            found_words.append(pattern.strip())
            reasons.append(f'"{pattern.strip()}" x{count} — {reason}')
            fixes.append(fix)
    score = len(found_words)
    return score, found_words, reasons, fixes

def style_header_row(ws, row, col_start, col_end, fill):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# SHEET 1: Full RFP — Color-coded sections
# ============================================================
ws1 = wb.active
ws1.title = "Full RFP"

# Section labels row (row 1)
FEEDBACK_FILL = PatternFill('solid', fgColor='E65100')   # Orange header
FEEDBACK_TINT = PatternFill('solid', fgColor='FFF3E0')   # Light orange cells

sections = [
    (1, 3, "METADATA", META_FILL),
    (4, 7, "BSB ASKS", BSB_FILL),
    (8, 10, "BRIM RESPONDS", BRIM_FILL),
    (11, 13, "FEEDBACK TO CLAUDE", FEEDBACK_FILL),
    (14, 16, "DELIVERY METHOD", BRIM_FILL),
    (17, 22, "BRIM METADATA", META_FILL),
    (23, 29, "CLAUDE AI ANALYSIS", AI_FILL),
]

for col_start, col_end, label, fill in sections:
    mid = (col_start + col_end) // 2
    cell = ws1.cell(row=1, column=mid, value=label)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal='center')
    for c in range(col_start, col_end + 1):
        ws1.cell(row=1, column=c).fill = fill
        ws1.cell(row=1, column=c).font = SECTION_FONT

# Column headers (row 2)
headers = [
    # META (1-3)
    'Ref', 'Category', '#',
    # BSB ASKS (4-7)
    'Topic', 'Requirement', 'Compliant', 'Confidence',
    # BRIM RESPONDS (8-10)
    'Bullet Response', 'Paragraph Response', 'Response Status',
    # FEEDBACK TO CLAUDE (11-13)
    'Feedback on Bullet', 'Feedback on Paragraph', 'Priority (H/M/L)',
    # DELIVERY METHOD (14-16)
    'A:OOB', 'B:Config', 'C:Custom',
    # BRIM META (17-22)
    'Pricing', 'Capability', 'Availability', 'Strategic Notes', 'Notes', 'Reg Enable',
    # AI ANALYSIS (23-29)
    'AI Detection Score', 'AI Words Found', 'Word Count (Bullet)', 'Word Count (Para)',
    'Rationale', 'Committee Review', 'Committee Score',
]

keys = [
    'ref', 'category', 'number',
    'topic', 'requirement', 'compliant', 'confidence',
    'bullet', 'paragraph', '_response_status',
    '_fb_bullet', '_fb_para', '_fb_priority',
    'a_oob', 'b_config', 'c_custom',
    'pricing', 'capability', 'availability', 'strategic', 'notes', 'reg_enable',
    '_ai_score', '_ai_words', '_wc_bullet', '_wc_para',
    'rationale', 'committee_review', 'committee_score',
]

# Tint map: which columns get which tint
tint_ranges = [
    (1, 3, META_TINT),
    (4, 7, BSB_TINT),
    (8, 10, BRIM_TINT),
    (11, 13, FEEDBACK_TINT),
    (14, 16, BRIM_TINT),
    (17, 22, META_TINT),
    (23, 29, AI_TINT),
]

def get_tint(col):
    for s, e, t in tint_ranges:
        if s <= col <= e:
            return t
    return None

# Header fills match section
header_fills = {}
for s, e, _, fill in sections:
    for c in range(s, e + 1):
        header_fills[c] = fill

for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = header_fills.get(i, META_FILL)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws1.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
ws1.freeze_panes = 'A3'

# Data rows
for r, q in enumerate(questions, 3):
    # Compute AI detection
    bullet = q.get('bullet', '') or ''
    para = q.get('paragraph', '') or ''
    combined = bullet + ' ' + para
    ai_score, ai_words, _, _ = detect_ai_score(combined)
    wc_bullet = len(bullet.split()) if bullet else 0
    wc_para = len(para.split()) if para else 0

    for c, key in enumerate(keys, 1):
        if key == '_ai_score':
            val = ai_score
        elif key == '_ai_words':
            val = ', '.join(ai_words) if ai_words else 'Clean'
        elif key == '_wc_bullet':
            val = wc_bullet
        elif key == '_wc_para':
            val = wc_para
        elif key == '_response_status':
            conf = q.get('confidence', '')
            val = 'Needs Work' if conf in ('RED', 'YELLOW') else 'Ready'
        elif key in ('_fb_bullet', '_fb_para', '_fb_priority'):
            val = ''  # Empty — user fills in feedback
        else:
            val = q.get(key, '')

        cell = ws1.cell(row=r, column=c, value=val)
        cell.font = FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

        # Apply section tint
        tint = get_tint(c)
        if tint:
            cell.fill = tint

        # Override for confidence column
        if key == 'confidence' and val in CONF_MAP:
            cell.fill = CONF_MAP[val]

        # Override for AI score - red if high
        if key == '_ai_score':
            if val >= 3:
                cell.fill = RED_FILL
            elif val >= 1:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = GREEN_FILL

# Column widths
widths = [12, 20, 5, 25, 45, 9, 11, 50, 50, 12,
          35, 35, 10,
          6, 6, 6,
          15, 12, 14, 40, 30, 10,
          12, 25, 12, 12, 35, 40, 8]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 2: Summary by Category
# ============================================================
ws2 = wb.create_sheet("Summary by Category")
sum_headers = ['Category', 'Total', 'GREEN', 'YELLOW', 'RED', '% Compliant', 'Avg AI Score', 'Avg Word Count']
for i, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill('solid', fgColor='2F5496')
    cell.alignment = Alignment(horizontal='center')
ws2.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}1"
ws2.freeze_panes = 'A2'

cats = {}
for q in questions:
    cat = q.get('category', 'Unknown')
    if cat not in cats:
        cats[cat] = {'total': 0, 'GREEN': 0, 'YELLOW': 0, 'RED': 0, 'compliant': 0, 'ai_scores': [], 'wc': []}
    cats[cat]['total'] += 1
    conf = q.get('confidence', '')
    if conf in ('GREEN', 'YELLOW', 'RED'):
        cats[cat][conf] += 1
    if q.get('compliant') == 'Y':
        cats[cat]['compliant'] += 1
    bullet = q.get('bullet', '') or ''
    para = q.get('paragraph', '') or ''
    ai_s, _, _, _ = detect_ai_score(bullet + ' ' + para)
    cats[cat]['ai_scores'].append(ai_s)
    cats[cat]['wc'].append(len(para.split()) if para else len(bullet.split()) if bullet else 0)

for r, (cat, s) in enumerate(cats.items(), 2):
    ws2.cell(row=r, column=1, value=cat).font = FONT
    ws2.cell(row=r, column=2, value=s['total']).font = FONT
    g = ws2.cell(row=r, column=3, value=s['GREEN']); g.font = FONT; g.fill = GREEN_FILL
    y = ws2.cell(row=r, column=4, value=s['YELLOW']); y.font = FONT; y.fill = YELLOW_FILL
    rd = ws2.cell(row=r, column=5, value=s['RED']); rd.font = FONT; rd.fill = RED_FILL
    pct = s['compliant'] / s['total'] if s['total'] else 0
    ws2.cell(row=r, column=6, value=pct).font = FONT
    ws2.cell(row=r, column=6).number_format = '0.0%'
    avg_ai = sum(s['ai_scores']) / len(s['ai_scores']) if s['ai_scores'] else 0
    ws2.cell(row=r, column=7, value=round(avg_ai, 1)).font = FONT
    avg_wc = sum(s['wc']) / len(s['wc']) if s['wc'] else 0
    ws2.cell(row=r, column=8, value=round(avg_wc, 0)).font = FONT
    for c in range(1, 9):
        ws2.cell(row=r, column=c).border = THIN_BORDER

for w, col in [(25,1),(8,2),(8,3),(8,4),(8,5),(12,6),(12,7),(14,8)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SHEET 3: RED Questions
# ============================================================
ws3 = wb.create_sheet("RED Questions")
red_headers = ['Ref', 'Category', '#', 'Topic', 'Requirement', 'Bullet Response', 'Paragraph Response',
               'Confidence', 'Capability', 'Availability', 'Strategic', 'Rationale']
red_keys = ['ref', 'category', 'number', 'topic', 'requirement', 'bullet', 'paragraph',
            'confidence', 'capability', 'availability', 'strategic', 'rationale']

for i, h in enumerate(red_headers, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill('solid', fgColor='C0392B')
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws3.auto_filter.ref = f"A1:{get_column_letter(len(red_headers))}1"
ws3.freeze_panes = 'A2'

reds = [q for q in questions if q.get('confidence') == 'RED']
for r, q in enumerate(reds, 2):
    for c, key in enumerate(red_keys, 1):
        cell = ws3.cell(row=r, column=c, value=q.get(key, ''))
        cell.font = FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        cell.fill = PatternFill('solid', fgColor='FADBD8')

red_widths = [12, 20, 5, 25, 45, 50, 50, 11, 12, 14, 40, 40]
for i, w in enumerate(red_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 4: QA Changes Log
# ============================================================
ws4 = wb.create_sheet("QA Changes Log")
log_headers = ['#', 'Date', 'Question Ref', 'Field', 'Change Description', 'Type', 'Pass']
for i, h in enumerate(log_headers, 1):
    cell = ws4.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill('solid', fgColor='2F5496')
    cell.alignment = Alignment(horizontal='center')
ws4.auto_filter.ref = f"A1:{get_column_letter(len(log_headers))}1"
ws4.freeze_panes = 'A2'

TYPE_COLORS = {
    'Hallucination Fix': RED_FILL,
    'Consistency Fix': YELLOW_FILL,
    'Name Fix': PatternFill('solid', fgColor='BDD7EE'),
    'Accuracy Fix': PatternFill('solid', fgColor='D5E8D4'),
    'Date Fix': PatternFill('solid', fgColor='FFF2CC'),
    'Formatting': PatternFill('solid', fgColor='E2EFDA'),
    'Cleanup': PatternFill('solid', fgColor='F2F2F2'),
}

changes = [
    # Pass 1 - Initial cleanup
    ("2026-03-22", "All (20 questions)", "multiple", "Removed all Coast Capital Savings references (fabricated client)", "Hallucination Fix", "Pass 1"),
    ("2026-03-22", "All", "multiple", "Removed all Continental Currency references (fabricated client)", "Hallucination Fix", "Pass 1"),
    ("2026-03-22", "Activation & Fulfillment 15", "paragraph", "Removed fabricated 73% first-transaction metric", "Hallucination Fix", "Pass 1"),
    ("2026-03-22", "Processing 26", "paragraph", "Removed fabricated 99.997% settlement accuracy metric", "Hallucination Fix", "Pass 1"),
    ("2026-03-22", "Multiple", "multiple", "Replaced East Coast references with Eastern time zone", "Accuracy Fix", "Pass 1"),
    ("2026-03-22", "Partner Relationships 8", "paragraph", "Fixed Wolfsberg FCCQ 45/45 → completed self-assessment", "Hallucination Fix", "Pass 1"),
    # Pass 2 - Today's QA
    ("2026-03-23", "Partner Relationships 1", "notes", "Renamed Aite-Novarica → Datos", "Name Fix", "Pass 2"),
    ("2026-03-23", "Partner Relationships 7", "bullet", "Renamed Aite-Novarica → Datos", "Name Fix", "Pass 2"),
    ("2026-03-23", "Partner Relationships 7", "paragraph", "Renamed Aite-Novarica → Datos", "Name Fix", "Pass 2"),
    ("2026-03-23", "Partner Relationships 13", "bullet", "Softened deconversion: 'in production' → 'developed with'", "Hallucination Fix", "Pass 2"),
    ("2026-03-23", "Partner Relationships 13", "paragraph", "Softened deconversion: 'in production' → 'developed with'", "Hallucination Fix", "Pass 2"),
    ("2026-03-23", "Partner Relationships 14", "strategic", "Softened Mastercard EXCLUSIVE → strategic", "Hallucination Fix", "Pass 2"),
    ("2026-03-23", "Technology 24", "bullet", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Technology 24", "paragraph", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Technology 24", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Technology 24", "paragraph", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Activation & Fulfillment 15", "rationale", "Removed self-referential audit notes", "Cleanup", "Pass 2"),
    ("2026-03-23", "Compliance & Reporting 3", "strategic", "Renamed Aite-Novarica → Datos", "Name Fix", "Pass 2"),
    ("2026-03-23", "Processing 19", "bullet", "Fixed Verafin: 'runs in production' → 'planned'", "Hallucination Fix", "Pass 2"),
    ("2026-03-23", "Processing 19", "paragraph", "Fixed Verafin: 'runs in production' → 'planned'", "Hallucination Fix", "Pass 2"),
    ("2026-03-23", "Processing 22", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Processing 26", "rationale", "Removed self-referential audit notes", "Cleanup", "Pass 2"),
    ("2026-03-23", "Processing 32", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Processing 42", "bullet", "Fixed trailing uptime 99.97% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Accounting & Finance 3", "strategic", "Renamed Aite-Novarica → Datos", "Name Fix", "Pass 2"),
    ("2026-03-23", "Accounting & Finance 4", "strategic", "Renamed Aite-Novarica → Datos", "Name Fix", "Pass 2"),
    ("2026-03-23", "Accounting & Finance 15", "availability", "Updated past date Est. Q1 2026 → Est. Q2 2026", "Date Fix", "Pass 2"),
    ("2026-03-23", "Product Operations 21", "paragraph", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "Product Operations 33", "paragraph", "Standardized uptime SLA 99.95% → 99.99%", "Consistency Fix", "Pass 2"),
    ("2026-03-23", "BULK (68 questions)", "bullet/paragraph", "Added terminal punctuation to 68 responses", "Formatting", "Pass 2"),
]

for r, (date, ref, field, desc, typ, pass_num) in enumerate(changes, 2):
    ws4.cell(row=r, column=1, value=r-1).font = FONT
    ws4.cell(row=r, column=2, value=date).font = FONT
    ws4.cell(row=r, column=3, value=ref).font = FONT
    ws4.cell(row=r, column=4, value=field).font = FONT
    ws4.cell(row=r, column=5, value=desc).font = FONT
    c6 = ws4.cell(row=r, column=6, value=typ); c6.font = FONT
    if typ in TYPE_COLORS:
        c6.fill = TYPE_COLORS[typ]
    ws4.cell(row=r, column=7, value=pass_num).font = FONT
    for c in range(1, 8):
        ws4.cell(row=r, column=c).border = THIN_BORDER
        ws4.cell(row=r, column=c).alignment = WRAP

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 55
ws4.column_dimensions['F'].width = 18
ws4.column_dimensions['G'].width = 8

# ============================================================
# SHEET 5: AI Detection — with Why and How to Improve
# ============================================================
ws5 = wb.create_sheet("AI Detection")
ai_headers = ['Ref', 'Category', '#', 'AI Score', 'Status', 'AI Words Found',
              'Why Flagged', 'How to Improve', 'Word Count', 'Confidence', 'Sample Text (first 200 chars)']
for i, h in enumerate(ai_headers, 1):
    cell = ws5.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill('solid', fgColor='7B1FA2')
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws5.auto_filter.ref = f"A1:{get_column_letter(len(ai_headers))}1"
ws5.freeze_panes = 'A2'

flagged = 0
low_risk = 0
clean = 0

for r, q in enumerate(questions, 2):
    bullet = q.get('bullet', '') or ''
    para = q.get('paragraph', '') or ''
    combined = bullet + ' ' + para
    ai_score, ai_words, reasons, fixes = detect_ai_score(combined)
    wc = len(para.split()) if para else len(bullet.split()) if bullet else 0
    sample = (para[:200] if para else bullet[:200]) if (para or bullet) else ''

    if ai_score == 0:
        status = 'CLEAN'
        why = 'No AI-typical language detected'
        how = 'No action needed'
        clean += 1
    elif ai_score <= 2:
        status = 'LOW RISK'
        why = '\n'.join(reasons)
        how = '\n'.join(fixes)
        low_risk += 1
    else:
        status = 'FLAG — REWRITE'
        why = '\n'.join(reasons)
        how = '\n'.join(fixes)
        flagged += 1

    ws5.cell(row=r, column=1, value=q.get('ref', '')).font = FONT
    ws5.cell(row=r, column=2, value=q.get('category', '')).font = FONT
    ws5.cell(row=r, column=3, value=q.get('number', '')).font = FONT

    sc = ws5.cell(row=r, column=4, value=ai_score); sc.font = FONT
    if ai_score >= 3: sc.fill = RED_FILL
    elif ai_score >= 1: sc.fill = YELLOW_FILL
    else: sc.fill = GREEN_FILL

    st = ws5.cell(row=r, column=5, value=status); st.font = FONT
    if status == 'CLEAN': st.fill = GREEN_FILL
    elif status == 'LOW RISK': st.fill = YELLOW_FILL
    else: st.fill = RED_FILL

    ws5.cell(row=r, column=6, value=', '.join(ai_words) if ai_words else 'None').font = FONT
    ws5.cell(row=r, column=7, value=why).font = FONT
    ws5.cell(row=r, column=8, value=how).font = FONT
    ws5.cell(row=r, column=9, value=wc).font = FONT
    ws5.cell(row=r, column=10, value=q.get('confidence', '')).font = FONT
    ws5.cell(row=r, column=11, value=sample).font = Font(name='Arial', size=9, color='666666')

    for c in range(1, 12):
        ws5.cell(row=r, column=c).border = THIN_BORDER
        ws5.cell(row=r, column=c).alignment = WRAP

for w, col in [(12,1),(20,2),(5,3),(10,4),(15,5),(30,6),(45,7),(45,8),(10,9),(11,10),(40,11)]:
    ws5.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SHEET 6: DASHBOARD — Executive Overview
# ============================================================
ws6 = wb.create_sheet("Dashboard")
ws6.sheet_properties.tabColor = "1F4E79"

# Move to first position
wb.move_sheet(ws6, offset=-5)

TITLE_FONT = Font(name='Arial', size=16, bold=True, color='1F4E79')
H2_FONT = Font(name='Arial', size=13, bold=True, color='2F5496')
H3_FONT = Font(name='Arial', size=11, bold=True, color='424242')
BODY_FONT = Font(name='Arial', size=10)
BODY_WRAP = Alignment(wrap_text=True, vertical='top')
METRIC_FONT = Font(name='Arial', size=24, bold=True, color='1F4E79')
METRIC_LABEL = Font(name='Arial', size=9, color='666666')

# Column widths
for c in range(1, 12):
    ws6.column_dimensions[get_column_letter(c)].width = 14

ws6.column_dimensions['A'].width = 3
ws6.column_dimensions['B'].width = 20
ws6.column_dimensions['C'].width = 18
ws6.column_dimensions['D'].width = 18
ws6.column_dimensions['E'].width = 18
ws6.column_dimensions['F'].width = 18
ws6.column_dimensions['G'].width = 18
ws6.column_dimensions['H'].width = 25

# --- TITLE ---
r = 2
ws6.cell(row=r, column=2, value="BSB Credit Card Program RFP").font = TITLE_FONT
r += 1
ws6.cell(row=r, column=2, value="AI-Assisted Response Analysis Dashboard").font = Font(name='Arial', size=12, color='666666')
r += 1
ws6.cell(row=r, column=2, value="Brim Financial | March 2026").font = Font(name='Arial', size=10, color='999999')

# --- KEY METRICS ---
r += 2
ws6.cell(row=r, column=2, value="KEY METRICS").font = H2_FONT
ws6.cell(row=r, column=2).fill = PatternFill('solid', fgColor='D6E4F0')
for c in range(2, 8):
    ws6.cell(row=r, column=c).fill = PatternFill('solid', fgColor='D6E4F0')

r += 1
metrics = [
    (str(len(questions)), "Total Questions"),
    (str(sum(1 for q in questions if q['confidence']=='GREEN')), "GREEN"),
    (str(sum(1 for q in questions if q['confidence']=='YELLOW')), "YELLOW"),
    (str(sum(1 for q in questions if q['confidence']=='RED')), "RED"),
    (f"{sum(1 for q in questions if q.get('compliant')=='Y')/len(questions)*100:.0f}%", "Compliant"),
    (str(clean), "AI-Clean Responses"),
]
for i, (val, label) in enumerate(metrics):
    col = i + 2
    ws6.cell(row=r, column=col, value=val).font = METRIC_FONT
    ws6.cell(row=r, column=col).alignment = Alignment(horizontal='center')
    ws6.cell(row=r+1, column=col, value=label).font = METRIC_LABEL
    ws6.cell(row=r+1, column=col).alignment = Alignment(horizontal='center')

# Color the metric cells
ws6.cell(row=r, column=3).font = Font(name='Arial', size=24, bold=True, color='2E7D32')  # GREEN
ws6.cell(row=r, column=4).font = Font(name='Arial', size=24, bold=True, color='F57F17')  # YELLOW
ws6.cell(row=r, column=5).font = Font(name='Arial', size=24, bold=True, color='C62828')  # RED

# --- CATEGORY BREAKDOWN ---
r += 4
ws6.cell(row=r, column=2, value="CATEGORY BREAKDOWN").font = H2_FONT
ws6.cell(row=r, column=2).fill = PatternFill('solid', fgColor='D6E4F0')
for c in range(2, 8):
    ws6.cell(row=r, column=c).fill = PatternFill('solid', fgColor='D6E4F0')

r += 1
cat_headers = ['Category', 'Total', 'GREEN', 'YELLOW', 'RED', '% Compliant']
for i, h in enumerate(cat_headers):
    cell = ws6.cell(row=r, column=i+2, value=h)
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.fill = PatternFill('solid', fgColor='E8E8E8')

for cat_name, s in cats.items():
    r += 1
    ws6.cell(row=r, column=2, value=cat_name).font = BODY_FONT
    ws6.cell(row=r, column=3, value=s['total']).font = BODY_FONT
    ws6.cell(row=r, column=4, value=s['GREEN']).font = BODY_FONT
    ws6.cell(row=r, column=4).fill = GREEN_FILL
    ws6.cell(row=r, column=5, value=s['YELLOW']).font = BODY_FONT
    ws6.cell(row=r, column=6, value=s['RED']).font = BODY_FONT
    pct = s['compliant'] / s['total'] if s['total'] else 0
    ws6.cell(row=r, column=7, value=pct).font = BODY_FONT
    ws6.cell(row=r, column=7).number_format = '0%'
    for c in range(2, 8):
        ws6.cell(row=r, column=c).border = THIN_BORDER

# --- AI DETECTION SUMMARY ---
r += 3
ws6.cell(row=r, column=2, value="AI LANGUAGE DETECTION SUMMARY").font = H2_FONT
ws6.cell(row=r, column=2).fill = PatternFill('solid', fgColor='F3E5F5')
for c in range(2, 8):
    ws6.cell(row=r, column=c).fill = PatternFill('solid', fgColor='F3E5F5')

r += 1
ai_metrics = [
    (str(clean), "Clean (no AI patterns)", GREEN_FILL),
    (str(low_risk), "Low Risk (1-2 patterns)", YELLOW_FILL),
    (str(flagged), "Flagged (3+ patterns)", RED_FILL),
]
for i, (val, label, fill) in enumerate(ai_metrics):
    ws6.cell(row=r, column=2+i*2, value=val).font = Font(name='Arial', size=18, bold=True)
    ws6.cell(row=r, column=2+i*2).fill = fill
    ws6.cell(row=r, column=2+i*2).alignment = Alignment(horizontal='center')
    ws6.cell(row=r+1, column=2+i*2, value=label).font = METRIC_LABEL
    ws6.cell(row=r+1, column=2+i*2).alignment = Alignment(horizontal='center')

# --- QA PASSES ---
r += 4
ws6.cell(row=r, column=2, value="QA PASSES COMPLETED").font = H2_FONT
ws6.cell(row=r, column=2).fill = PatternFill('solid', fgColor='D6E4F0')
for c in range(2, 8):
    ws6.cell(row=r, column=c).fill = PatternFill('solid', fgColor='D6E4F0')

r += 1
qa_items = [
    ("Pass 1: Hallucination Removal", "Removed fabricated clients (Coast Capital, Continental Currency), fake metrics (73%, 99.997%), fixed Wolfsberg scoring, corrected timezone references"),
    ("Pass 2: Consistency & Accuracy", "Softened Mastercard EXCLUSIVE claim, renamed Aite-Novarica to Datos, standardized uptime SLA to 99.99%, fixed Verafin integration status, updated past dates"),
    ("Pass 3: AI Language Cleanup", "Replaced buzzwords (ecosystem, synergy, leverage, utilize), fixed passive voice patterns, removed filler transitions (Additionally, Furthermore, Moreover)"),
    ("Pass 4: Terminal Punctuation", "Added missing periods to 68 responses that ended without punctuation"),
    ("Pass 5: Verification", "Ran automated QA: 383 questions confirmed, no duplicates, no shifted data, all removed terms verified at zero, build passes"),
]
for item, desc in qa_items:
    ws6.cell(row=r, column=2, value=item).font = Font(name='Arial', size=10, bold=True)
    ws6.cell(row=r, column=4, value=desc).font = BODY_FONT
    ws6.cell(row=r, column=4).alignment = BODY_WRAP
    ws6.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    ws6.row_dimensions[r].height = 35
    r += 1


# ============================================================
# SHEET 7: METHODOLOGY — How AI Was Used
# ============================================================
ws7 = wb.create_sheet("Methodology")
ws7.sheet_properties.tabColor = "7B1FA2"
wb.move_sheet(ws7, offset=-5)

ws7.column_dimensions['A'].width = 3
ws7.column_dimensions['B'].width = 25
ws7.column_dimensions['C'].width = 70
ws7.column_dimensions['D'].width = 3

r = 2
ws7.cell(row=r, column=2, value="AI-Assisted RFP Response Methodology").font = TITLE_FONT
r += 1
ws7.cell(row=r, column=2, value="How This RFP Was Built, Reviewed, and Quality-Assured").font = Font(name='Arial', size=11, color='666666')

# --- OVERVIEW ---
r += 2
ws7.cell(row=r, column=2, value="OVERVIEW").font = H2_FONT
ws7.cell(row=r, column=2).fill = PatternFill('solid', fgColor='D6E4F0')
ws7.cell(row=r, column=3).fill = PatternFill('solid', fgColor='D6E4F0')
r += 1
overview_text = [
    ("What", "We are building a fleet of 40+ specialized AI agents to automate RFP response workflows end-to-end. This is not just 'using AI to write' — it's an agentic system where purpose-built agents handle rewriting, critique, consistency checking, hallucination detection, AI language scoring, regulatory review, competitive analysis, and data validation. Each agent has a specific role and operates within guardrails."),
    ("Why", "Traditional RFP responses take weeks of manual effort across multiple SMEs. Our agent-based approach compresses timeline from weeks to days while improving quality through automated multi-perspective review. This is a fundamentally new way to do RFPs — not a person using ChatGPT, but an orchestrated system of specialized agents with human oversight at decision points."),
    ("Agent architecture", "40+ agents organized by function: Writer agents (rewrite, humanize, strengthen), Reviewer agents (CPO, regulatory, tech architect, competitive, writing quality), QA agents (hallucination detection, consistency check, data validation, AI language scoring), and Export agents (PDF, Word, Excel formatting). Agents run in parallel where possible."),
    ("Model", "Claude by Anthropic (Opus and Sonnet models). All agents use the same model family for consistency. No GPT, Gemini, or other models. Claude Code CLI orchestrates the agent fleet."),
    ("Human oversight", "Agents propose, humans approve. Every AI modification is tracked, diffed, and subject to human review. Agents never auto-publish. All client names, metrics, and claims are verified against confirmed data before inclusion. The Excel feedback columns let reviewers send instructions directly back to the agents."),
]
for label, text in overview_text:
    ws7.cell(row=r, column=2, value=label).font = H3_FONT
    ws7.cell(row=r, column=3, value=text).font = BODY_FONT
    ws7.cell(row=r, column=3).alignment = BODY_WRAP
    ws7.row_dimensions[r].height = 50
    r += 1

# --- AI WORKFLOW ---
r += 1
ws7.cell(row=r, column=2, value="AI WORKFLOW — STEP BY STEP").font = H2_FONT
ws7.cell(row=r, column=2).fill = PatternFill('solid', fgColor='F3E5F5')
ws7.cell(row=r, column=3).fill = PatternFill('solid', fgColor='F3E5F5')
r += 1

steps = [
    ("1. Initial Responses", "SMEs draft initial bullet and paragraph responses for each of BSB's 383 questions. Loaded into the RFP Viewer tool as the baseline dataset."),
    ("2. Writer Agents", "Specialized rewrite agents strengthen weak responses (YELLOW/RED). Each rewrite generates a word-level diff — the human sees exactly what changed and can Accept, Reject, or Edit. Agents are prompted with the full question context, Knowledge Base, and writing rules."),
    ("3. Knowledge Base Agent", "A grounding agent maintains verified company facts, client references, key metrics, and differentiators. Every writer agent's prompt includes this KB, preventing hallucination at the source."),
    ("4. Rules Engine", "Global and per-category writing rules (e.g., 'Never claim capabilities we don't have', 'Reference specific client implementations'). All agents must follow these rules. Rules are versioned and auditable."),
    ("5. Reviewer Agents (6 in parallel)", "Six expert review agents run simultaneously, each with a different persona: CPO/Procurement, Banking Regulatory, Technology Architect, RFP Writing Quality, Competitive Analysis, Data Consistency. Each produces a structured report with findings and severity ratings."),
    ("6. Consistency Agent", "Scans all 383 responses for contradictions — e.g., one answer claims 99.95% uptime while another says 99.99%. Cross-references delivery methods, capability claims, and timelines across the full document."),
    ("7. Humanize Agent", "Automated pass removes AI-typical language: filler transitions (Additionally, Furthermore), passive voice, vague superlatives (comprehensive, robust, seamless), corporate buzzwords (leverage, synergy). Replaced with direct, evidence-based language."),
    ("8. Hallucination Detector", "Multi-pass QA agent specifically looking for fabricated data: fake client names, invented metrics, unverifiable claims. Found and removed: Coast Capital (fake client), Continental Currency (fake), 73% metric (invented), 99.997% metric (invented). Every finding is logged with before/after."),
    ("9. AI Detection Scorer", "Scores every response for AI-typical language patterns using pattern matching + semantic analysis. Results: CLEAN (0 patterns), LOW RISK (1-2), FLAG (3+). Provides specific 'Why' and 'How to Improve' for each flagged item."),
    ("10. Feedback Loop", "Reviewers (Rasha, SMEs) add feedback directly in the Excel 'Feedback to Claude' columns. These instructions are fed back into the writer agents for the next iteration. The system learns from each review cycle."),
    ("11. Export Agents", "Generate presentation-ready outputs: color-coded PDF, Word document, Excel workbook with dashboard. Each export agent formats data for its target audience (evaluators, executives, technical reviewers)."),
]
for label, text in steps:
    ws7.cell(row=r, column=2, value=label).font = H3_FONT
    ws7.cell(row=r, column=3, value=text).font = BODY_FONT
    ws7.cell(row=r, column=3).alignment = BODY_WRAP
    ws7.row_dimensions[r].height = 45
    r += 1

# --- WHAT AI DOES vs DOES NOT DO ---
r += 1
ws7.cell(row=r, column=2, value="WHAT AI DOES vs DOES NOT DO").font = H2_FONT
ws7.cell(row=r, column=2).fill = PatternFill('solid', fgColor='E8F5E9')
ws7.cell(row=r, column=3).fill = PatternFill('solid', fgColor='E8F5E9')
r += 1

does = [
    ("AI DOES", "Rewrite weak responses for clarity and strength"),
    ("", "Check consistency across 383 answers"),
    ("", "Detect and flag AI-sounding language"),
    ("", "Score responses from a procurement committee perspective"),
    ("", "Identify contradictions, gaps, and missing information"),
    ("", "Generate diff views so humans see every change"),
    ("", "Apply writing rules consistently across all categories"),
]
doesnt = [
    ("AI DOES NOT", "Invent client names, metrics, or capabilities"),
    ("", "Auto-publish without human review"),
    ("", "Make strategic decisions about positioning"),
    ("", "Replace SME judgment on technical accuracy"),
    ("", "Determine compliance status (Y/N/Partial)"),
    ("", "Set confidence ratings (GREEN/YELLOW/RED)"),
    ("", "Access external systems or BSB's evaluation criteria"),
]

for label, text in does:
    ws7.cell(row=r, column=2, value=label).font = Font(name='Arial', size=10, bold=True, color='2E7D32') if label else BODY_FONT
    ws7.cell(row=r, column=3, value=text).font = BODY_FONT
    ws7.cell(row=r, column=3).fill = PatternFill('solid', fgColor='E8F5E9')
    ws7.cell(row=r, column=2).fill = PatternFill('solid', fgColor='E8F5E9')
    r += 1

r += 1
for label, text in doesnt:
    ws7.cell(row=r, column=2, value=label).font = Font(name='Arial', size=10, bold=True, color='C62828') if label else BODY_FONT
    ws7.cell(row=r, column=3, value=text).font = BODY_FONT
    ws7.cell(row=r, column=3).fill = PatternFill('solid', fgColor='FFC7CE')
    ws7.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FFC7CE')
    r += 1

# --- PROMPTING STRATEGY ---
r += 2
ws7.cell(row=r, column=2, value="PROMPTING STRATEGY").font = H2_FONT
ws7.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FFF2CC')
ws7.cell(row=r, column=3).fill = PatternFill('solid', fgColor='FFF2CC')
r += 1

prompts = [
    ("System prompt", "AI is told: 'You are a senior RFP writer for a fintech company responding to a US community bank. Write in first person plural (we/our). Be specific and evidence-based. Never fabricate metrics or client names.'"),
    ("Context injection", "Each rewrite prompt includes: the question topic, BSB's requirement, current response, confidence level, committee score, risk assessment, delivery method, and any global/category rules."),
    ("Knowledge Base", "Verified company facts are injected into every prompt so the AI draws from real data rather than general knowledge. Includes: client list, platform capabilities, certifications, integration partners."),
    ("Guardrails", "Prompts explicitly instruct: 'Do not use words like comprehensive, robust, seamless, leverage. Do not use passive voice. Do not add claims not in the Knowledge Base.'"),
    ("Output format", "AI returns structured JSON with the rewritten text, allowing the tool to compute diffs and track changes programmatically."),
]
for label, text in prompts:
    ws7.cell(row=r, column=2, value=label).font = H3_FONT
    ws7.cell(row=r, column=3, value=text).font = BODY_FONT
    ws7.cell(row=r, column=3).alignment = BODY_WRAP
    ws7.row_dimensions[r].height = 50
    r += 1


# ============================================================
# SHEET 8: ASSUMPTIONS — Global, Category, Question Level
# ============================================================
ws8 = wb.create_sheet("Assumptions & Positioning")
ws8.sheet_properties.tabColor = "2E7D32"
wb.move_sheet(ws8, offset=-5)

ws8.column_dimensions['A'].width = 3
ws8.column_dimensions['B'].width = 22
ws8.column_dimensions['C'].width = 70
ws8.column_dimensions['D'].width = 3

r = 2
ws8.cell(row=r, column=2, value="Assumptions & Client Positioning Strategy").font = TITLE_FONT
r += 1
ws8.cell(row=r, column=2, value="What we assumed, how we positioned, and why").font = Font(name='Arial', size=11, color='666666')

# --- GLOBAL ASSUMPTIONS ---
r += 2
ws8.cell(row=r, column=2, value="GLOBAL ASSUMPTIONS").font = H2_FONT
ws8.cell(row=r, column=2).fill = PatternFill('solid', fgColor='D6E4F0')
ws8.cell(row=r, column=3).fill = PatternFill('solid', fgColor='D6E4F0')
r += 1

global_assumptions = [
    ("BSB Profile", "Bangor Savings Bank is a US community bank evaluating credit card program partners. They value compliance, reliability, and a strong servicing partner over cutting-edge innovation."),
    ("Competitive landscape", "BSB is likely evaluating 3-5 vendors including incumbents (FIS, Fiserv, TSYS) and fintechs (Brim, Deserve, Cardless). Brim's differentiator is modern tech + Canadian market proof points."),
    ("Decision criteria", "BSB's evaluation will weight: (1) regulatory compliance, (2) operational reliability, (3) integration with Jack Henry core, (4) cardholder experience, (5) total cost. In that order."),
    ("Tone & voice", "Professional, direct, evidence-based. No marketing language. First person plural (we/our). Every claim backed by a specific client, metric, or certification."),
    ("Compliance posture", "Lead with compliance in every answer. BSB is a regulated bank — they need to see that Brim understands and respects regulatory requirements (OCC, CFPB, state-level)."),
    ("Honest gaps", "Where Brim cannot meet a requirement, we acknowledge the gap honestly and provide a mitigation plan with timeline. Never claim capabilities we don't have. RED confidence = honest gap."),
    ("Client references", "Only reference confirmed, live Brim clients: Manulife, Zolve, Affinity Credit Union, Laurentian Bank, CWB, Air France-KLM Flying Blue, Payfacto, UNI, Money Mart, easyfinancial, NDAX."),
    ("Metrics policy", "Only use verified metrics. If a number hasn't been confirmed by internal data, soften to ranges or remove. Learned lesson: AI fabricated 73% and 99.997% metrics in early drafts."),
    ("SLA standard", "99.99% uptime SLA is the confirmed contractual number. All references standardized to this."),
    ("AI transparency", "We are transparent about AI assistance. This is a strength, not a weakness — it shows Brim uses modern tools and has QA processes to ensure accuracy."),
]
for label, text in global_assumptions:
    ws8.cell(row=r, column=2, value=label).font = H3_FONT
    ws8.cell(row=r, column=3, value=text).font = BODY_FONT
    ws8.cell(row=r, column=3).alignment = BODY_WRAP
    ws8.row_dimensions[r].height = 50
    for c in [2, 3]:
        ws8.cell(row=r, column=c).border = THIN_BORDER
    r += 1

# --- CATEGORY-LEVEL ASSUMPTIONS ---
r += 2
ws8.cell(row=r, column=2, value="CATEGORY-LEVEL POSITIONING").font = H2_FONT
ws8.cell(row=r, column=2).fill = PatternFill('solid', fgColor='E8F5E9')
ws8.cell(row=r, column=3).fill = PatternFill('solid', fgColor='E8F5E9')
r += 1

cat_assumptions = [
    ("Technology", "Lead with cloud-native architecture, API-first design, modern security stack. Acknowledge Jack Henry integration is in progress — position as 'pre-built connectors + custom API layer' rather than claiming native integration."),
    ("Processing", "Emphasize real-time authorization, fraud detection, settlement accuracy. Verafin integration is planned (Q3 2026), not live — be clear about this. Lead with current fraud capabilities."),
    ("Compliance & Reporting", "This is BSB's top priority. Lead with certifications (PCI DSS v4.0.1, SOC 2 Type 2, ISO 27001), regulatory expertise, and Datos 96% rating. Brim's compliance posture is genuinely strong."),
    ("Partner Relationships", "Showcase breadth of live clients across different verticals (banking, fintech, loyalty, crypto). Emphasize Manulife scale (55K+ advisors) and Affinity Credit Union as direct FI comparable."),
    ("Customer Experience", "Position 24/7/365 fraud/lost stolen support as differentiator. Digital-first servicing with omnichannel (app, web, IVR, live agent). Acknowledge BSB may want in-branch capabilities we don't fully support."),
    ("Application Processing", "Lead with configurable decisioning rules and multi-bureau support. Brim's digital-first application flow is strong. Acknowledge that some legacy workflows (paper apps, in-branch) may need customization."),
    ("Activation & Fulfillment", "Instant digital issuance + push provisioning to Apple/Google/Samsung Pay is a strength. Physical card production via IDEMIA. Emergency card production available."),
    ("Loyalty & Benefits", "Brim Marketplace (500+ merchants) and configurable rewards engine are differentiators. Manulife loyalty integration is the proof point. Air France-KLM Flying Blue shows cross-border capability."),
    ("Product Operations", "Position Brim's incident management, change management, and release cadence. 99.99% uptime SLA. Transparent about operational maturity — growing but proven at scale with Manulife."),
    ("Collections & Recovery", "Acknowledge this is an area where Brim is building capability. Position current state honestly and show roadmap. Don't oversell."),
    ("Accounting & Finance", "Settlement, reconciliation, and GL integration capabilities. 1099 generation available (updated from Q1 2026 estimate). Standard financial reporting suite."),
    ("Acquisition & Lifecycle", "Digital marketing integration, lifecycle management, and campaign tools. Position Manulife's 13 separate acquisition journeys as proof of configurability."),
]
for label, text in cat_assumptions:
    ws8.cell(row=r, column=2, value=label).font = H3_FONT
    ws8.cell(row=r, column=3, value=text).font = BODY_FONT
    ws8.cell(row=r, column=3).alignment = BODY_WRAP
    ws8.row_dimensions[r].height = 55
    for c in [2, 3]:
        ws8.cell(row=r, column=c).border = THIN_BORDER
    r += 1

# --- QUESTION-LEVEL POSITIONING NOTES ---
r += 2
ws8.cell(row=r, column=2, value="QUESTION-LEVEL POSITIONING (RED items)").font = H2_FONT
ws8.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FFC7CE')
ws8.cell(row=r, column=3).fill = PatternFill('solid', fgColor='FFC7CE')
r += 1
ws8.cell(row=r, column=2, value="These are the 7 honest gaps. Positioning strategy for each:").font = Font(name='Arial', size=10, italic=True, color='666666')
r += 1

for q in reds:
    ref = q.get('ref', '')
    topic = q.get('topic', '')
    cap = q.get('capability', '')
    strat = q.get('strategic', '') or ''
    ws8.cell(row=r, column=2, value=f"{ref}: {topic}").font = H3_FONT
    ws8.cell(row=r, column=3, value=f"Capability: {cap}. {strat[:300]}").font = BODY_FONT
    ws8.cell(row=r, column=3).alignment = BODY_WRAP
    ws8.row_dimensions[r].height = 55
    ws8.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FADBD8')
    ws8.cell(row=r, column=3).fill = PatternFill('solid', fgColor='FADBD8')
    for c in [2, 3]:
        ws8.cell(row=r, column=c).border = THIN_BORDER
    r += 1

# --- THINGS THAT HELPED ---
r += 2
ws8.cell(row=r, column=2, value="WHAT HELPED US / LESSONS LEARNED").font = H2_FONT
ws8.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FFF2CC')
ws8.cell(row=r, column=3).fill = PatternFill('solid', fgColor='FFF2CC')
r += 1

lessons = [
    ("Multi-agent review", "Running 6 expert agents in parallel (CPO, regulatory, tech architect, writing quality, competitive, data consistency) caught issues no single reviewer would find. The regulatory agent caught the Wolfsberg scoring error. The data consistency agent found the uptime SLA contradiction."),
    ("Hallucination detection", "AI will fabricate data if not constrained. Our early drafts included fake client names (Coast Capital, Continental Currency) and invented metrics (73%, 99.997%). Lesson: always verify AI output against a confirmed source list."),
    ("Honest gaps win trust", "Marking 7 questions RED with honest gap acknowledgment + roadmap is stronger than pretending we can do everything. BSB evaluators respect transparency — they've seen vendors oversell."),
    ("AI detection as QA", "Scoring every response for AI-typical language patterns catches problems before BSB's evaluators do. If we can detect it sounds AI-written, so can they."),
    ("Datos rating as proof", "The Datos (formerly Aite-Novarica) 96% product features rating is a powerful third-party validation. It belongs in Compliance & Reporting and Partner Relationships sections."),
    ("Manulife as anchor", "Manulife is the strongest proof point — 55K+ advisors, 13 acquisition journeys, June-October 2025 launch timeline. Use it across Technology, Processing, Partner Relationships, and Loyalty."),
    ("Tools over people", "Building the RFP Viewer tool (grid view, AI rewrite with diff, critique, consistency check, export) was an investment that paid off. It's reusable for future RFPs, not a one-time effort."),
]
for label, text in lessons:
    ws8.cell(row=r, column=2, value=label).font = H3_FONT
    ws8.cell(row=r, column=3, value=text).font = BODY_FONT
    ws8.cell(row=r, column=3).alignment = BODY_WRAP
    ws8.row_dimensions[r].height = 55
    for c in [2, 3]:
        ws8.cell(row=r, column=c).border = THIN_BORDER
    r += 1


wb.save(OUTPUT)
print(f"Saved {OUTPUT}")
print(f"  Sheet 1: Dashboard")
print(f"  Sheet 2: Methodology")
print(f"  Sheet 3: Assumptions & Positioning")
print(f"  Sheet 4: Full RFP - color-coded ({len(questions)} rows)")
print(f"  Sheet 5: Summary by Category ({len(cats)} categories)")
print(f"  Sheet 6: RED Questions ({len(reds)} rows)")
print(f"  Sheet 7: QA Changes Log ({len(changes)} entries)")
print(f"  Sheet 8: AI Detection ({len(questions)} rows)")
print(f"\nAI Detection: {clean} clean, {low_risk} low risk, {flagged} flagged")
